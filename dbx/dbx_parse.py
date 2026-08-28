"""Defensive parser for the feasible_database.xlsx P_<pair> lattice sheets.

Read-only.  Produces a python dict per pair; every field records whether it was
found, so a sheet that deviates from the modal layout degrades to a partial
record instead of raising.
"""
from __future__ import annotations

import re
import sys
from pathlib import Path

import openpyxl

ROOT = Path(r"c:\Users\USER\Desktop\CT&RPL\2_Project\KNF_LEU+ 시범연료봉 장전 인허가 과제\2026\2_계산\5_RL")
XLSX = ROOT / "data" / "reports" / "scoping_mesh_20260815" / "feasible_database.xlsx"
sys.path.insert(0, str(ROOT))

from lpopt.data.fuel_types import (  # noqa: E402
    kconv_curve_shape,
    _interp,
    _burnup_at_k1,
)

# Korean item labels in the [일반 / 핀 농축도 / 독봉] block -> canonical key.
ITEM_MAP = {
    "집합체 평균 농축도 (w/o)": "u_avg_enrichment",
    "역할 (경/중부하)": "role",
    "주연료 U-235 (w/o)": "enr_main",
    "zoning 저농축 U-235 (w/o)": "enr_zone",
    "zoning 격차 du (w/o)": "du",
    "독봉(Gd) 개수": "n_gd",
    "독봉 Gd2O3 (wt%)": "gd_wt",
    "독봉 U-235 (w/o)": "gd_u_enr",
    "Gd 배치 패턴": "gd_pattern",
    "예측 FF (연소창 최대)": "ff_window_max",
}
NUMERIC_ITEMS = {
    "u_avg_enrichment", "enr_main", "enr_zone", "du",
    "n_gd", "gd_wt", "gd_u_enr", "ff_window_max",
}
GT_CODES = {6, 7, 8, 9}


def _s(v):
    return "" if v is None else str(v).strip()


def _num(v):
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def load_rows(ws):
    return [list(r) for r in ws.iter_rows(values_only=True)]


def octant_counts(tri: list[list[int]]) -> dict[int, int] | None:
    """Expand a 1/8 (lower-triangular, diagonal-symmetric) pin map to full counts.

    Row i (0-based) holds j = 0..i.  Diagonal cells (i == j) have multiplicity 4,
    off-diagonal 8 -> 8*4 + 28*8 = 256 for an 8-row (16x16) octant.
    """
    counts: dict[int, int] = {}
    for i, row in enumerate(tri):
        if len(row) != i + 1:
            return None
        for j, code in enumerate(row):
            w = 4 if i == j else 8
            counts[code] = counts.get(code, 0) + w
    return counts


def parse_sheet(ws) -> dict:
    rows = load_rows(ws)
    rec: dict = {"sheet": ws.title, "problems": [], "blocks": []}

    # ---- title -----------------------------------------------------------
    m = re.match(r"집합체 쌍\s+(\S+)\s*\(세그먼트\s*([\d.]+)", _s(rows[0][0]))
    if m:
        rec["pair"] = m.group(1)
        rec["segment"] = float(m.group(2))
    else:
        rec["pair"] = ws.title[2:]
        rec["problems"].append("title row unparsed")

    # ---- locate blocks ---------------------------------------------------
    marks = {}
    for i, r in enumerate(rows):
        c0 = _s(r[0])
        if c0.startswith("["):
            marks.setdefault(c0.split("]")[0] + "]", i)
            rec["blocks"].append(c0)

    # ---- general block ---------------------------------------------------
    types: list[str] = []
    hdr = next((i for i, r in enumerate(rows) if _s(r[0]) == "항목"), None)
    if hdr is None:
        rec["problems"].append("no 항목 header")
    else:
        types = [_s(c) for c in rows[hdr][1:] if _s(c)]
        rec["types"] = types
        gen = {t: {} for t in types}
        found = set()
        for r in rows[hdr + 1:]:
            label = _s(r[0])
            if not label or label.startswith("["):
                break
            key = ITEM_MAP.get(label)
            if key is None:
                rec["problems"].append(f"unmapped item label {label!r}")
                continue
            found.add(key)
            for k, t in enumerate(types):
                raw = r[1 + k] if 1 + k < len(r) else None
                val = _num(raw) if key in NUMERIC_ITEMS else (_s(raw) or None)
                if val is not None:
                    gen[t][key] = val
        rec["general"] = gen
        missing_items = set(ITEM_MAP.values()) - found
        if missing_items:
            rec["problems"].append(f"item rows absent: {sorted(missing_items)}")

    # ---- feed distribution ----------------------------------------------
    fi = next((i for i, r in enumerate(rows) if _s(r[0]) == "feed"), None)
    if fi is None:
        rec["problems"].append("no feed-distribution header")
    else:
        dist = []
        for r in rows[fi + 1:]:
            f = _num(r[0])
            if f is None:
                break
            dist.append({"feed": int(f), "n_pass": int(_num(r[1]) or 0),
                         "best_f_r": _num(r[2])})
        rec["feed_dist"] = dist

    # ---- pin maps --------------------------------------------------------
    pinmaps = {}
    for i, r in enumerate(rows):
        mm = re.match(r"\[1/8 핀맵 — (\S+)\]", _s(r[0]))
        if not mm:
            continue
        tid = mm.group(1)
        entry = {"deck": None, "tri": None, "legend": {}, "counts": None}
        j = i + 1
        if _s(rows[j][0]).startswith("(덱 없음"):
            entry["deck"] = False
            pinmaps[tid] = entry
            continue
        entry["deck"] = True
        dm = re.search(r"\((\S+\.inp)\)", _s(r[0])) or re.search(r"\(DeCART (\S+)\)", _s(r[0]))
        entry["deck_file"] = dm.group(1) if dm else None
        tri: list[list[int]] = []
        while j < len(rows):
            r2 = rows[j]
            c0, c1 = _num(r2[0]), _num(r2[1]) if len(r2) > 1 else None
            if c0 is None:
                break
            # a legend row is "code | <text>"; a map row is all-numeric
            if len(r2) > 1 and _s(r2[1]) and c1 is None:
                break
            vals = [int(_num(c)) for c in r2 if _num(c) is not None]
            tri.append(vals)
            j += 1
        entry["tri"] = tri
        # legend
        while j < len(rows):
            r2 = rows[j]
            code = _num(r2[0])
            desc = _s(r2[1]) if len(r2) > 1 else ""
            if code is None or not desc:
                break
            entry["legend"][int(code)] = desc
            j += 1
        counts = octant_counts(tri)
        if counts is None:
            entry["problem"] = f"non-triangular pin map ({[len(x) for x in tri]})"
            rec["problems"].append(f"{tid}: " + entry["problem"])
        else:
            entry["counts"] = counts
            entry["n_positions"] = sum(counts.values())
        pinmaps[tid] = entry
    rec["pinmaps"] = pinmaps

    # ---- k-inf / FF curves ----------------------------------------------
    ki = next((i for i, r in enumerate(rows)
               if _s(r[0]).startswith("burnup")), None)
    if ki is None:
        rec["problems"].append("no k-inf table header")
    else:
        head = [_s(c) for c in rows[ki]]
        cols = {}
        for k, h in enumerate(head):
            if h.startswith("k_"):
                cols.setdefault("k", {})[h[2:]] = k
            elif h.startswith("FF_"):
                cols.setdefault("ff", {})[h[3:]] = k
        bu, kv, ffv = [], {}, {}
        for r in rows[ki + 1:]:
            b = _num(r[0])
            if b is None:
                break
            bu.append(b)
            for t, c in cols.get("k", {}).items():
                kv.setdefault(t, []).append(_num(r[c]))
            for t, c in cols.get("ff", {}).items():
                ffv.setdefault(t, []).append(_num(r[c]))
        rec["bu"] = bu
        rec["kinf"] = kv
        rec["ff"] = ffv
        if types and set(kv) != set(types):
            rec["problems"].append(
                f"k-inf columns {sorted(kv)} != types {sorted(types)}")
    return rec


def derive(rec: dict) -> dict:
    """Per-type derived lattice + curve features (never raises)."""
    out = {}
    for t in rec.get("types", []):
        d = dict(rec.get("general", {}).get(t, {}))
        d["pair"] = rec["pair"]
        d["segment"] = rec.get("segment")
        pm = rec.get("pinmaps", {}).get(t, {})
        d["has_deck"] = bool(pm.get("deck"))
        d["deck_file"] = pm.get("deck_file")
        cnt = pm.get("counts")
        if cnt:
            d["n_positions"] = sum(cnt.values())
            d["n_guide_tube"] = sum(v for k, v in cnt.items() if k in GT_CODES)
            d["n_fuel_rod"] = d["n_positions"] - d["n_guide_tube"]
            d["n_main_pin"] = cnt.get(1, 0)
            d["zone_pin_count"] = cnt.get(2, 0)
            d["n_gd_from_map"] = cnt.get(3, 0)
            # legend-recovered enrichments (authoritative when the item block is blank)
            for code, desc in pm.get("legend", {}).items():
                mm = re.search(r"주연료\s*([\d.]+)\s*w/o", desc)
                if mm:
                    d.setdefault("enr_main", float(mm.group(1)))
                mm = re.search(r"zoning 저농축\s*([\d.]+)\s*w/o", desc)
                if mm:
                    d.setdefault("enr_zone", float(mm.group(1)))
                mm = re.search(r"Gd봉\s*U([\d.]+)\s*w/o\s*\+\s*Gd2O3\s*([\d.]+)\s*wt%", desc)
                if mm:
                    d.setdefault("gd_u_enr", float(mm.group(1)))
                    d.setdefault("gd_wt", float(mm.group(2)))
            if all(k in d for k in ("enr_main", "enr_zone", "gd_u_enr")) and d["n_fuel_rod"]:
                d["u_avg_from_map"] = (
                    d["n_main_pin"] * d["enr_main"]
                    + d["zone_pin_count"] * d["enr_zone"]
                    + d["n_gd_from_map"] * d["gd_u_enr"]
                ) / d["n_fuel_rod"]
            if "du" not in d and "enr_main" in d and "enr_zone" in d:
                d["du"] = round(d["enr_main"] - d["enr_zone"], 4)
            if "n_gd" not in d and "n_gd_from_map" in d:
                d["n_gd"] = float(d["n_gd_from_map"])
        bu = rec.get("bu")
        ys = rec.get("kinf", {}).get(t)
        if bu and ys and len(bu) == len(ys) and all(y is not None for y in ys):
            d["kinf_n_pts"] = len(bu)
            d["kinf_bu_max"] = bu[-1]
            for g, col in zip((0.0, 10.0, 20.0, 30.0),
                              ("kinf0", "kinf10", "kinf20", "kinf30")):
                v = _interp(bu, ys, g)
                if v is not None:
                    d[col] = v
            bk = _burnup_at_k1(bu, ys)
            if bk is not None:
                d["bu_k1"] = bk
            d.update(kconv_curve_shape(bu, ys))
        ffs = rec.get("ff", {}).get(t)
        if ffs and all(f is not None for f in ffs):
            d["ff_pin_max"] = max(ffs)
            d["ff_boc"] = ffs[0]
        out[t] = d
    return out


def parse_all():
    wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
    recs, types = {}, {}
    for name in wb.sheetnames:
        if not name.startswith("P_"):
            continue
        rec = parse_sheet(wb[name])
        rec["derived"] = derive(rec)
        recs[rec["pair"]] = rec
        types.update(rec["derived"])
    wb.close()
    return recs, types


if __name__ == "__main__":
    recs, types = parse_all()
    print(f"{len(recs)} P_ sheets, {len(types)} assembly types")
    for p, r in recs.items():
        print(f"  {p:8s} types={r.get('types')} deck="
              f"{[t for t, v in r['pinmaps'].items() if v.get('deck')]} "
              f"kpts={len(r.get('bu', []))} problems={r['problems']}")
